from flask import Flask,jsonify
from flask import render_template, request, redirect
from flask import Flask, request, Response, render_template, send_file
from flask import Flask,  render_template, request, redirect, url_for, session # pip install Flask
from flask_mysqldb import MySQL,MySQLdb # pip install Flask-MySQLdb
from os import path #pip install notify-py
from openpyxl import Workbook
#librerias para cancer bucal..
import pickle
import numpy as np
import io
import os
from PIL import Image
from tensorflow.keras.preprocessing import image

app = Flask(__name__,template_folder='template')
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'sistema1'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'
mysql = MySQL(app)

#subiendo el archivo de cancer bucal generado en jupyter
phish_model_cancer = pickle.load(open(r'c:\Users\GABRIEL\Pictures\Tesis Gabriel\PROTOTIPO DE TESIS\tesis phising\tesis cancer\modelo_cancer_bucal.pkl', 'rb'))

def cancerPrediction(path):
    # Loading Image
    img = image.load_img(path, target_size=(256, 256))
    # Normalizing Image
    norm_img = image.img_to_array(img) / 255
    # Converting Image to Numpy Array
    input_arr_img = np.array([norm_img])
    # Getting Predictions
    pred = (phish_model_cancer.predict(input_arr_img) > 0.5).astype(int)[0][0]
    # Return Model Prediction
    if pred == 0:
       print("Cancer")
     
    else:
        print("No tiene Cancer")
    
   
    #return "Cáncer bucal" if pred == 0 else "No es cáncer bucal"

@app.route('/cancer_predict', methods=['POST'])
def cancer_predict():
    # Cargar el modelo de detección de cáncer bucal
    phish_model_cancer = pickle.load(open(r'c:\Users\GABRIEL\Pictures\Tesis Gabriel\PROTOTIPO DE TESIS\tesis phising\tesis phising\modelo_cancer_bucal.pkl', 'rb'))
    
    # Obtener la imagen enviada por el cliente
    file = request.files['file']

    # Verificar si el campo de archivo está vacío
    if file.filename == '':
        mensajes="Por favor seleecione una imagen antes de proceder"
        return render_template("cancer.html", sms=mensajes)

    # Obtener la ruta temporal para guardar la imagen
    temp_path = "temp_img.jpg"
    file.save(temp_path)

    # Realizar la predicción utilizando el modelo cargado
    # Loading Image
    img = image.load_img(temp_path, target_size=(256, 256))
    # Normalizing Image
    norm_img = image.img_to_array(img) / 255
    # Converting Image to Numpy Array
    input_arr_img = np.array([norm_img])
    # Getting Predictions
    pred = (phish_model_cancer.predict(input_arr_img) > 0.5).astype(int)[0][0]
    
    # Eliminar la imagen temporal
    os.remove(temp_path)
    
    # Return Model Prediction
    if pred == 0:
        print("Cancer")
        resu = "Cancer"
    else:
        print("No tiene Cancer")
        resu = "No tiene Cancer"

    # Ejemplo de respuesta (ajusta esto según la estructura de tu modelo y tus necesidades)
    return render_template("cancer.html", pred=resu)


@app.route('/cancer')
def cancer():
    return render_template('cancer.html')


# subiendo el archivo picle pkl generado en jupyter
phish_model_ls = pickle.load(open(r'c:\Users\GABRIEL\Pictures\Tesis Gabriel\PROTOTIPO DE TESIS\tesis phising\tesis phising\ataques.pkl', 'rb'))

urlError = {
    "Por favor ingrese el campo de URL"
}
#####################
@app.route('/')
def home():
    return render_template('index.html')    

@app.route('/layout-static')
def stati():
    return render_template("layout-static.html")  

@app.route('/layout-sidenav')
def sidenav():
    return render_template("layout-sidenav-light.html")  

@app.route('/contenido')
def inicio():
    return render_template("contenido.html")    
   

@app.route('/layout', methods = ["GET", "POST"])
def layout():
    session.clear()
    return render_template("contenido.html")

@app.route('/login')
def logueo():
    return render_template('login.html')
#----------- LOGIN --------------------------
@app.route('/acceso-login', methods= ["GET", "POST"])
def login():
    actividad = """SELECT usuarios.nombre, sitios_analizados.descripcion, sitios_analizados.url FROM sitios_analizados JOIN usuarios ON sitios_analizados.id_us = usuarios.id WHERE usuarios.id = %s"""
    if request.method == 'POST' and 'txtCorreo' in request.form and 'txtPassword' in request.form:
       
        _correo = request.form['txtCorreo']
        _password = request.form['txtPassword']

        cur = mysql.connection.cursor()
        cur.execute('SELECT * FROM usuarios WHERE correo = %s AND password = %s', (_correo, _password,))
        account = cur.fetchone()
      
        if account:
            session['logueado'] = True
            session['id'] = account['id']
            session['nombre'] = account['nombre']
            session['correo'] = account['correo']
            session['password'] = account['password']
            session['id_rol'] = account['id_rol']
            mensaje5="bienvenido!"

            if session['id_rol'] == 1:
                    return render_template("admin.html")
            elif session['id_rol'] == 2:
                    return render_template("usuario.html",mensajeinicio=mensaje5)
            print(account[0])
        else:
         
            return render_template('login.html',mensaje="Usuario O Contraseña Incorrectas")
  
    return render_template('login.html')
#---REGISTRO DE USUARIOS------------
@app.route('/registro')
def reg():
    return render_template('registro.html')

@app.route('/registro-crear', methods = ["GET", "POST"])
def registro():
    
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM roles")
    roles = cur.fetchall()
    cur.close()


    if request.method == 'GET':
        return render_template("registro.html", tipo = roles )
    
    else:
        name = request.form['txtNombre']
        email = request.form['txtCorreo']
        password = request.form['txtPassword']

        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM usuarios WHERE correo=%s",(email,))
        result = cur.fetchone()
        if result:
            return render_template("registro.html",resultado="El Usuario Ya Existe")
        
        cur.execute("INSERT INTO usuarios (nombre, correo, password, id_rol) VALUES (%s,%s,%s,'2')", (name, email, password))
        mysql.connection.commit()
        
        return render_template('login.html',mensaje2="Registro Exitoso")
    
#ELIMINAR DATOS DE LA TABLA DE USUARIOS ADMIN
@app.route('/eliminar/<int:id>')
def eliminar(id):
    cur = mysql.connection.cursor()
     
    cur.execute("DELETE FROM usuarios WHERE id=%s", (id,))
    mysql.connection.commit()
    
    return redirect('../us_admin')


#----PREDICCION--PAGINA--FORM
@app.route('/prediccion')
def prediccion():
  
    return render_template('prediccion.html')

@app.route('/estadisticas', methods = ["GET", "POST"])
def estadistica():
     cur = mysql.connection.cursor()
     
     querySQL = ("SELECT * FROM sitios")
     cur.execute(querySQL)
     data = cur.fetchall()
     cur.close()
     cur.close()
     return render_template('estadisticas.html', dataEmpleados = data )

@app.route('/admin', methods = ["GET", "POST"])
def administrador():
     cur = mysql.connection.cursor()
     
     querySQL = ("SELECT * FROM usuarios")
     cur.execute(querySQL)
     dato = cur.fetchall()
     cur.close()
     #si no esta logueado que lo bote a login
     if not session.get('logueado'):
        return redirect('login')
    ###############################
     return render_template('admin.html', datoEmpleados = dato )
 
#USUARIOS IMPRESOS EN EL ADMINISTRADOR TODOS LOS USUARIOS A EDITAR 
@app.route('/us_admin', methods = ["GET", "POST"])
def us_admin():
     user_id=session.get('id')
     
     cur = mysql.connection.cursor()
     
     querySQL = ("SELECT usuarios.*, roles.nom_rol AS rol FROM usuarios JOIN roles ON usuarios.id_rol = roles.id_rol")
     cur.execute(querySQL)
     dato = cur.fetchall()
     cur.close()
      #si no esta logueado que lo bote a login
     if not session.get('logueado'):
        return redirect('login')
    ###############################
     return render_template('usuarios_admin.html', datoEmpleados = dato )
 
#CUESTIONARIO REALIZADO POR USUARIOS 
@app.route('/charts', methods = ["GET", "POST"])
def cuest():
     cur = mysql.connection.cursor()
     
     querySQL = ("SELECT * FROM cuestionario")
     cur.execute(querySQL)
     dato = cur.fetchall()
     cur.close()
     
     return render_template('charts.html', cuestionario1 = dato )

@app.route('/charts')
def chart():
    #si no esta logueado que lo bote a login
    if not session.get('logueado'):
        return redirect('login')
    ###############################
    return render_template("charts.html")   

@app.route('/tabs', methods = ["GET", "POST"])
def urls():
     cur = mysql.connection.cursor()
     
     querySQL = ("SELECT * FROM sitios")
     cur.execute(querySQL)
     dat_ur = cur.fetchall()
     cur.close()
     return render_template('tables.html', datos_url = dat_ur )

@app.route('/actividad', methods = ["GET", "POST"])
def actividad():
     
     cur = mysql.connection.cursor()
     querySQL = """SELECT usuarios.nombre, sitios_analizados.descripcion, sitios_analizados.url FROM sitios_analizados JOIN usuarios ON sitios_analizados.id_us = usuarios.id WHERE usuarios.id = %s"""
     id_usuario = 22
     cur.execute(querySQL,(id_usuario,))
     acti = cur.fetchall()
     cur.close()
     
     return render_template('usuario.html', actividad = acti)
 
@app.route('/sit_an', methods = ["GET", "POST"])
def sit_an():
     cur = mysql.connection.cursor()
     
     querySQL = ("SELECT * FROM sitios_analizados")
     cur.execute(querySQL)
     dat_ur = cur.fetchall()
     cur.close()
     return render_template('admin_sit_an.html', datos_url = dat_ur )
 
#----PREDICCION--CON--MODELO ENTRENADO-----ATAQUES.PKL
@app.route('/predict',  methods=['POST'])
def predict():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM sitios_analizados")
    roles = cur.fetchall()
    cur.close()
    
    X_predict = []

    url = request.form.get("EnterYourSite")
    print(url, "0000000000000000000000")
    if url:
        X_predict.append(str(url))
        y_Predict = ''.join(phish_model_ls.predict(X_predict))
        print(y_Predict)
        if y_Predict == 'bad':
            result = "Este Es Un Sitio PHISHING ⛔"
            name = request.form['EnterYourSite']
            cur = mysql.connection.cursor()
            cur.execute("INSERT INTO `sitios_analizados` (`id`, `url`, `descripcion`) VALUES (NULL, %s, 'Phishing')", [name])
            mysql.connection.commit()
            print("sitio agregado")
        else:
            result = "Este Es Un Sitio LEGITIMO ✅"
            name = request.form['EnterYourSite']
            cur = mysql.connection.cursor()
            cur.execute("INSERT INTO `sitios_analizados` (`id`, `url`, `descripcion`) VALUES (NULL, %s, 'Legitimo')", [name])
            mysql.connection.commit()
            print("sitio agregado")
        return render_template('prediccion.html', prediction_text = result)

    elif not url:
        return Response(
            response=urlError,
            status=400
        )
 
#----PREDICCION--CON--MODELO ENTRENADO--EN PAGINA SITIOS ANALIZADOS POR EL USUARIO-----      
@app.route('/analiz',  methods=['POST'])
def usua():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM sitios_analizados")
    roles = cur.fetchall()
    cur.close()
    id_usuario = session.get('id')
    
    X_predict = []

    url = request.form.get("EnterYourSite")
    print(url, "0000000000000000000000")
    if url:
        X_predict.append(str(url))
        y_Predict = ''.join(phish_model_ls.predict(X_predict))
        print(y_Predict)
        if y_Predict == 'bad':
            result = "Este Es Un Sitio PHISHING ⛔"
            name = request.form['EnterYourSite']
            cur = mysql.connection.cursor()
            cur.execute("INSERT INTO `sitios_analizados` (`id`, `url`, `descripcion`,`id_us`) VALUES (NULL, %s, 'Phishing',%s)", [name,id_usuario])
            mysql.connection.commit()
            print("sitio agregado")
        else:
            result = "Este Es Un Sitio LEGITIMO ✅"
            name = request.form['EnterYourSite']
            cur = mysql.connection.cursor()
            cur.execute("INSERT INTO `sitios_analizados` (`id`, `url`, `descripcion`,`id_us`) VALUES (NULL, %s, 'Legitimo',%s)", [name,id_usuario])
            mysql.connection.commit()
            print("sitio agregado")
        return render_template('analizar_sitio_us.html', prediction_text = result)
#

#PAGINA DE SITIOS ANALIZADOS
@app.route('/list_analiz', methods = ["GET", "POST"])
def mostrar_actividad():
    
     cur = mysql.connection.cursor()
     user_id=session.get('id')
     cur.execute("SELECT * FROM sitios_analizados WHERE id_us = %s", (user_id,))
     analizados = cur.fetchall()
     cur.close()
     return render_template('sitios_analizados_usuarios.html', datos_analiz = analizados )  
#GENERAR REPORTE DE SITIOS ANALIZADOS POR EL USUARIO REGISTRADO-- SOLO DEL USUARIO LOGUEADO
@app.route('/generar_reporte', methods = ["GET", "POST"])
def generar_reporte():
     cur = mysql.connection.cursor()
     user_id=session.get('id')
     cur.execute("SELECT url, descripcion, id_us FROM sitios_analizados WHERE id_us = %s", (user_id,))
     analizados = cur.fetchall()
     
     workbook = Workbook()
     sheet = workbook.active
     column_names = [column[0] for column in cur.description]
     sheet.append(column_names)
     for row in analizados:
        sheet.append(list(row.values()))

     reporte_path = "reporte.xlsx"
     workbook.save(reporte_path)
     cur.close()
     return send_file(reporte_path, as_attachment=True)
#GENERAR LISTA DE USUARIOS REPORTE
@app.route('/reporte_usuario', methods = ["GET", "POST"])
def reporte_usuario():
     cur = mysql.connection.cursor()
     user_id=session.get('id')
     cur.execute("SELECT usuarios.*, roles.nom_rol AS rol FROM usuarios JOIN roles ON usuarios.id_rol = roles.id_rol")
     analizados = cur.fetchall()
     
     workbook = Workbook()
     sheet = workbook.active
     column_names = [column[0] for column in cur.description]
     sheet.append(column_names)
     for row in analizados:
        sheet.append(list(row.values()))

     reporte_path = "reporte.xlsx"
     workbook.save(reporte_path)
     cur.close()
     return send_file(reporte_path, as_attachment=True)
 
#GENERAR REPORTE DE LOS SITIOS ANALIZADOS POR LOS USUARIOS-- MUESTRA DE TODOS LOS USUAIROS
@app.route('/reporte_sitios_analizados', methods = ["GET", "POST"])
def reporte_sitios():
     cur = mysql.connection.cursor()
     user_id=session.get('id')
     cur.execute("SELECT sitios_analizados.id, sitios_analizados.url, sitios_analizados.descripcion, sitios_analizados.id_us, usuarios.nombre FROM sitios_analizados JOIN usuarios ON sitios_analizados.id_us = usuarios.id")
     analizados = cur.fetchall()
     
     workbook = Workbook()
     sheet = workbook.active
     column_names = [column[0] for column in cur.description]
     sheet.append(column_names)
     for row in analizados:
        sheet.append(list(row.values()))

     reporte_path = "reporte.xlsx"
     workbook.save(reporte_path)
     cur.close()
     return send_file(reporte_path, as_attachment=True)
 
##GENERAR REPORTE DE LOS SITIOS ANALIZADOS POR TODOS LOS USUARIOS Y NO USUARIOS DE MANERA GENERAL 
@app.route('/reporte_sitios_general', methods = ["GET", "POST"])
def reporte_sitios_general():
     cur = mysql.connection.cursor()
     user_id=session.get('id')
     cur.execute("SELECT id,url,descripcion FROM sitios_analizados ")
     analizados = cur.fetchall()
     
     workbook = Workbook()
     sheet = workbook.active
     column_names = [column[0] for column in cur.description]
     sheet.append(column_names)
     for row in analizados:
        sheet.append(list(row.values()))

     reporte_path = "reporte.xlsx"
     workbook.save(reporte_path)
     cur.close()
     return send_file(reporte_path, as_attachment=True)
 
##GENERAR REPORTE DE CUESTIONARIO DE LOS USUARIOS ENCUESTADOS 
@app.route('/reporte_cuestionario', methods = ["GET", "POST"])
def reporte_cuestionario():
     cur = mysql.connection.cursor()
     user_id=session.get('id')
     cur.execute("SELECT * FROM cuestionario ")
     analizados = cur.fetchall()
     
     workbook = Workbook()
     sheet = workbook.active
     column_names = [column[0] for column in cur.description]
     sheet.append(column_names)
     for row in analizados:
        sheet.append(list(row.values()))

     reporte_path = "reporte.xlsx"
     workbook.save(reporte_path)
     cur.close()
     return send_file(reporte_path, as_attachment=True)
@app.route('/usuario')
def usuari():
    if not session.get('logueado'):
        return redirect('login')
    return render_template('usuario.html')  

@app.route('/list_analiz')
def list_analiz():
    #si no esta logueado que lo bote a login
    if not session.get('logueado'):
        return redirect('login')
    ###############################
    return render_template('sitios_analizados_usuarios.html')  

#CUESTIONARIO DEL USUARIO
@app.route('/cuestionario')
def cuestionario():
    #si no esta logueado que lo bote a login
     if not session.get('logueado'):
        return redirect('login')
    ###############################
     return render_template('cuestionario.html')
 
#CREACION DE CUESTIONARIO
@app.route('/crear-cuestionario', methods = ["GET", "POST"])
def preguntas():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM cuestionario")
    roles = cur.fetchall()
    cur.close()
    user_id=session.get('id')

    if request.method == 'GET':
        return render_template("cuestionario.html", tipo = roles )
    
    else:
        name = request.form['txtNombre']
        preg1 = request.form['txtPreg1']
        preg2 = request.form['txtPreg2']
        preg3 = request.form['txtPreg3']
        preg4 = request.form['txtPreg4']
       

        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO cuestionario (nombre, preg1, preg2, preg3, preg4,id_usu) VALUES (%s,%s,%s,%s,%s,%s)", (name, preg1, preg2, preg3, preg4,user_id))
        mysql.connection.commit()
        
        return render_template('cuestionario.html',mensajecuestionario="Cuestionario Realizado")
    
#MENSAJES DE LOS USUARIOS PUBLICOS
@app.route('/crear-mensajes', methods = ["GET", "POST"])
def crear_mensajes():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM mensajes")
    roles = cur.fetchall()
    cur.close()
    #user_id=session.get('id')
    if request.method == 'GET':
        return render_template("index.html", tipo = roles )
    
    else:
        correo = request.form['txtEmail']
        mensaje = request.form['txtMensaje']
        
        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO mensajes (correo, mensaje) VALUES (%s,%s)", (correo,mensaje))
        mysql.connection.commit()
        
        return render_template('index.html',mensajepublico="Mensaje Enviado Exitosamente")
#MOSTRAR MENSAJES--    
@app.route('/mensajes', methods = ["GET", "POST"])
def mostrar_mensajes():
    
     cur = mysql.connection.cursor()
     
     cur.execute("SELECT * FROM mensajes ")
     mensajes_admin = cur.fetchall()
     cur.close()
     return render_template('mensajes_admin.html', mensajes = mensajes_admin )  
 
#PAGINA DE MENSAJES
@app.route('/mensajes')
def mensajes_admin():
    #si no esta logueado que lo bote a login
    if not session.get('logueado'):
        return redirect('login')
    ###############################
    return render_template('mensaje_admin.html')
#####################

@app.route('/analiz')
def pg_analiz():
     if not session.get('logueado'):
            return redirect('login')
     return render_template('analizar_sitio_us.html')

#destruir sesion
@app.route('/cerrar-sesion')
def logout():
    session.clear()
    return render_template("login.html")

    
if __name__ == '__main__':
   app.secret_key = "pinchellave"
   app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)
